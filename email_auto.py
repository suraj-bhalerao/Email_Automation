import smtplib
import ssl
import time
import random
from email.message import EmailMessage
from tkinter import SE
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import shutil

load_dotenv()

SENDER_EMAIL = os.getenv("SENDER_EMAIL")
APP_PASSWORD = os.getenv("APP_PASSWORD")

EXCEL_PATH = os.getenv("EXCEL_PATH")
RESUME_PATH = os.getenv("RESUME_PATH")

if not SENDER_EMAIL:
    print("Error: SENDER_EMAIL environment variable is not set.")
    exit(1)
if not APP_PASSWORD:
    print("Error: APP_PASSWORD environment variable is not set.")
    exit(1)
if not EXCEL_PATH:
    print("Error: EXCEL_PATH environment variable is not set.")
    exit(1)
if not RESUME_PATH:
    print("Error: RESUME_PATH environment variable is not set.")
    exit(1)

DELAY_MIN = 15    
DELAY_MAX = 120    

SUBJECTS = [
    "Application for QA Automation Engineer | Java & Selenium",
    "Automation Test Engineer",
    "QA Automation Test Engineer",
    "Application for QA Automation Engineer Role",
]

BODIES = [
"""
Hi,

I hope you’re having a good day.

I am Suraj Bhalerao, and I’m a QA Automation Engineer  with around  1.6+ years of hands-on experience  in Java-based Selenium automation and API testing. I’m currently working with  Accolade Electronics Pvt Ltd , where I focus on building scalable automation frameworks and improving overall test reliability.

In my current role, I design  TestNG-based automation frameworks , automate end-to-end  UI regression suites, and validate backend APIs using  Postman . I’ve also integrated automation pipelines with  CI/CD (GitHub Actions)  and enjoy working on solutions that reduce manual effort and improve release confidence.

I’m actively exploring new opportunities where I can contribute as a QA Automation Engineer and continue growing in test automation and quality engineering. I believe my experience could be a good fit for teams that value clean automation design and strong testing fundamentals.

I’ve attached my resume for your reference.
You can also take a look at my work here:

LinkedIn: https://www.linkedin.com/in/suraj-bhalerao27
GitHub: https://github.com/suraj-bhalerao
Leetcode: https://leetcode.com/u/Suraj_b_27/

If there are any current or upcoming opportunities that match my profile, I’d be happy to connect and discuss further.

Thank you for your time and consideration.

Warm regards,
Suraj Bhalerao 
""",
"""
Hello,

I hope you’re doing well.

My name is Suraj Bhalerao, and I’m a QA Automation Engineer with 1.6+ years of hands-on experience in Selenium-based automation using Java, along with API testing experience. I’m currently working with Accolade Electronics Pvt Ltd, where I contribute to building robust automation frameworks and improving testing efficiency.

My experience includes developing TestNG automation frameworks, executing end-to-end UI regression automation, validating APIs through Postman, and integrating automation with CI/CD pipelines via GitHub Actions. I enjoy working on automation solutions that enhance product quality and reduce manual effort.

I’m now seeking opportunities where I can continue to grow as a QA Automation Engineer and add value to teams that prioritize automation and quality engineering.

My resume is attached for your review. You can also explore my profile here:

LinkedIn: https://www.linkedin.com/in/suraj-bhalerao27
GitHub: https://github.com/suraj-bhalerao
Leetcode: https://leetcode.com/u/Suraj_b_27/

I’d be glad to connect if there’s a suitable opportunity.

Kind regards,
Suraj Bhalerao
""",
"""
Hi,

I hope you’re having a great day.

I’m Suraj Bhalerao, a QA Automation Engineer with around 1.6+ years of experience working on Java-based Selenium automation and API testing. Currently, I work at Accolade Electronics Pvt Ltd, where I focus on creating scalable automation solutions and improving test reliability across projects.

My day-to-day work includes designing TestNG automation frameworks, automating complete UI regression flows, and validating backend APIs using Postman. I’ve also integrated automation runs into CI/CD pipelines with GitHub Actions, helping teams catch issues earlier and reduce repetitive manual testing.

I’m actively looking for new opportunities where I can apply my automation skills and continue learning in a quality-focused engineering environment. I enjoy building clean, maintainable frameworks and contributing to stable, well-tested releases.

I’ve shared my resume for your reference, and you can also check out my work below:

LinkedIn: https://www.linkedin.com/in/suraj-bhalerao27
GitHub: https://github.com/suraj-bhalerao
Leetcode: https://leetcode.com/u/Suraj_b_27/

If my profile seems relevant to any current or upcoming roles, I’d love to connect.

Thank you for your time.

Best regards,
Suraj Bhalerao
"""
]

context = ssl.create_default_context()

with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
    server.login(SENDER_EMAIL, APP_PASSWORD)

    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    
    if sheet is None:
        print("Error: No active sheet found in the workbook.")
        exit()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        to_email = row[0]

        subject = random.choice(SUBJECTS)
        body = random.choice(BODIES)

        msg = EmailMessage()
        msg["From"] = SENDER_EMAIL
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.set_content(body)

        with open(RESUME_PATH, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="pdf",
                filename="Suraj_Bhalerao_QA_Automation_Resume.pdf"
            )

        try:
            server.send_message(msg)
            print(f"Sent to {to_email} | Subject: {subject}")
        except Exception as e:
            print(f"Failed for {to_email}: {e}")

        delay = random.randint(DELAY_MIN, DELAY_MAX)
        time.sleep(delay)

shutil.move(EXCEL_PATH, os.path.join("isSent", os.path.basename(EXCEL_PATH)))
print(f"Moved {EXCEL_PATH} to isSent folder.")
